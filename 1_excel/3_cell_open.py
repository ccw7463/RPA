from openpyxl import load_workbook # 파일 불러오기 역할

excel = load_workbook("sample.xlsx") # EXCEL 불러오기

ws = excel.active # 활성화된 Sheet

# cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(x,y).value, end=" ") # print에 의한 자동 줄바꿈 방지하기위해 end 사용
    print() # 줄바꿈

print("-"*30)

# cell 개수를 모를 때 데이터 불러오기
'''
$ : worksheet
$.max_row : 최대 행 인덱스 리턴
$.max_column : 최대 열 인덱스 리턴
'''
for x in range(1,ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(x,y).value,end=" ")
    print()

excel.save("sample.xlsx")
excel.close()