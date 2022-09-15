# ▶▶▶ 셀 가져오기 심화 ◀◀◀
from openpyxl import load_workbook
excel = load_workbook("sample.xlsx")
ws = excel["score"]


# 영어 점수 열만 가져오기
'''
- ws["C"] : column B인 값들을 튜플형태로 전부 가져옴
- ws["A:C"] : column 이 A~C 인 값들을 전부 가져올수도 있음
'''
eng_score2 = []
eng_column = ws["C"]
for i in eng_column:
    eng_score2.append(i.value)
print("영어 점수2",eng_score2)


# 행 가져오기 + 각 셀에 대한 정보 가져오기
from openpyxl.utils.cell import coordinate_from_string # 정보 가져올때 필요
'''
- ws[1] : 1번째 row 가져오기
- ws[2:w1.max_row] : 2번째~마지막 row 가져오기
'''
row_title = ws[1] 
for cell in row_title:
    print(cell.value)

row_range = ws[2:ws.max_row]
for rows in row_range:
    for cell in rows:
        # print(cell.value, end=" ") # --> 셀 값 가져오기
        # print(cell.coordinate,end=" ") # --> 셀 정보 가져오기

        # 셀정보 슬라이싱해서 가져오기
        xy = coordinate_from_string(cell.coordinate)
        print(xy ,end=" ")
        print(xy[0],end=" ")
        print(xy[1])
        print("-"*15)
    print()


# ★ 중요 ★ --> 아래 2번 방법을 제일많이쓴다!
'''
전체/범위 행 가져오기
1. tuple(ws.rows)
2. ws.iter_rows()
    - 2번방식의 경우 범위 지정이가능해서좋음
    - ws.iter_rows(min_row=2,max_row=5)
    - ws.iter_rows(min_row=2,max_row=5,min_col=2,max_col=4)
    - 보통 for문에서 in뒤에서 사용한다.
전체/범위 열 가져오기
1. tuple(ws.columns)
2. ws.iter_cols()
'''
print(tuple(ws.rows)) # --> ((A1,B1,C1,D1),(A2,B2,C2,D2),...)
print(tuple(ws.columns)) # --> ((A1,A2,..,A11),(B1,B2,...,B11),..)

#  저장하기
excel.save("sample.xlsx")
excel.close()