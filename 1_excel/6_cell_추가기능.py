# ▶▶▶ 셀 추가 기능을 다룬다. ◀◀◀
# 1. 찾기 : 원하는 셀을 가져오기
# 2. 삽입 : 엑셀에서 지원하는 "삽입"기능 
# 3. 삭제
# 4. 이동

from openpyxl import load_workbook
excel = load_workbook("sample.xlsx")

# worksheet 확인 및 할당
print(excel.sheetnames)
ws = excel["score"]

'''
1. 찾기
'''
# 영어 점수 일정 점수 이상 학생 서치
for row in ws.iter_rows(min_row=2):
    # 번호, 국어, 영어, 수학
    if int(row[2].value) > 80:
        print(row[0].value, "번 학생 : 1등급")

'''
2. 삽입
'''
# ws.insert_rows(8) # 8번째 행이 비워짐
# ws.insert_rows(8,5) # 8번째 행위치부터 5줄을 삽입
# ws.insert_cols(2) # 2번째 열이 비워짐
# ws.insert_cols(2,2) # 2번째 열부터 2열을 삽입

'''
3. 삭제
'''
# ws.delete_rows(4) # 4번째 행이 지워짐
# ws.delete_rows(4,2) # 4번째 행부터 두개행이 지워짐
# ws.delete_cols(3) 
# ws.delete_cols(3,2)
# excel.save("sample_insert_rows.xlsx")
# excel.close()

'''
4. 이동
    - 현재 : 번호, 국어, 영어, 수학
    - 변경 : 번호, 국어, (과학), 영어, 수학
        --> 과학을 추가생성하기 위해, 이동수행 
    사용되는 메소드
    - move_range(범위,rows=값,cols=값)
        - 범위는 셀범위
        - rows와 cols 값은 양수/음수 가능
        - rows : 양수/음수 일때 아래/위 이동
        - cols : 양수/음수 일때 오른쪽/왼쪽 이동
'''
ws.move_range("C1:D11",rows=0,cols=1) # --> 행은 그대로, 열은 오른쪽 한칸 이동
ws.cell(1,3).value = "과학"
    # 다른방법1) ws["C1"] = "과학"
    # 다른방법2) ws.cell(1,3,"과학")


excel.save("modified.xlsx")
excel.close()

