# ▶▶▶ 셀 채우기 심화 ◀◀◀
from openpyxl import Workbook
excel = Workbook()
ws = excel.active

# 반복문을 이용해서 랜덤 숫자 채우기
from random import *
index = 1
for x in range(1,11): # 10 개 row
    for y in range(1,11): # 10 개 column
        # ws.cell(x,y,randint(0,100))
        ws.cell(x,y,index)
        index = index + 1

# 리스트로 셀 채우기
ws1 = excel.create_sheet("score")
ws1.sheet_properties.tabColor = "ff33cc"
ws1.append(["번호","국어","영어","수학"]) # 리스트로 넣음

for i in range(1,11): # 10개 데이터 넣기
    ws1.append([i,randint(70,100),randint(0,100),randint(0,100)])

excel.save("sample.xlsx")
excel.close()