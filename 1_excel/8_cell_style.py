from openpyxl import load_workbook
excel = load_workbook("sample.xlsx")
ws = excel["score"]

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1행의 높이를 50으로 설정 (인덱스 1부터 시작)
ws.row_dimensions[1].height = 50

# 스타일 적용
from openpyxl.styles import Font
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자색(빨강), 기울이기, 두껍게 설정
b1.font = Font(color="00FF00", name="Arial", strike=True) # 글자색(초록), 폰트변경, 취소선 설정
c1.font = Font(color="0000FF",size=15, underline="single") # 글자색(파랑), 크기15, 밑줄

# 테두리 적용
from openpyxl.styles import Border,Side
thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),bottom=Side(style="thin"),top=Side(style="thin"))
a1.border = thin_border


# 85 넘는 셀에 대해서 색칠
from openpyxl.styles import Alignment 
from openpyxl.styles import PatternFill
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center",vertical="center")
        if cell.column == 1: # 번호열은 제외
            continue
        # 셀이 정수형 데이터이고 85점보다 크면
        if isinstance(cell.value,int) and cell.value > 85 :
            cell.fill = PatternFill(fgColor="222222",fill_type="solid")
            cell.font = Font(color="FF0000")

# 틀 고정 --> "B2" 셀 기준으로 가로세로 틀 고정선 생김
ws.freeze_panes = "B2" 

# 저장 
excel.save("sample_style.xlsx")
excel.close()