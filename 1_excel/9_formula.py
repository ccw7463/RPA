import datetime
from openpyxl import Workbook
excel = Workbook()
ws = excel.active

ws.column_dimensions["A"].width = 15
ws["A1"] = datetime.date.today() # 오늘 날짜 정보
ws["A2"] = "=SUM(1,2,3,4)"
ws["A3"] = "=AVERAGE(1,2,3,4,5)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

excel.save("formula.xlsx")
# excel.close()

from openpyxl import load_workbook
excel = load_workbook("formula.xlsx")
ws = excel.active

# 수식 그대로 가져오기 (ws.iter_rows() 써도 동일)
for row in ws.values:
    for cell in row:
        print(cell)
print("-"*20)

# 보여지는 값을 가져오기
'''
보여지는 값을 그대로 가져오려면 load_workbook(이름,data_only=True)
None 으로 나타나는 값은 버그가 아니다
    - 파이썬으로 엑셀파일을 제어할때, 수식자체가 들어간거기 때문에 계산되지 않았어서 값이 안나옴
'''
excel = load_workbook("formula.xlsx",data_only=True)
ws = excel.active

for row in ws.values:
    for cell in row:
        print(cell)

