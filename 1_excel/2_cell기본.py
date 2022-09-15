'''
기본틀 1. 생성
from openpyxl import Workbook
excel = Workbook()
ws = excel.active
ws.title = "NadoSheet"

기본틀 2. 로드
from openpyxl import load_workbook
excel = load_workbook("sample.xlsx")
ws = excel.active()
'''

from openpyxl import Workbook
excel = Workbook()
ws = excel.active
ws.title = "NadoSheet"

# ▶▶▶ 기본 ◀◀◀
# 셀에 값넣기
'''
값 넣는법
1. ws["A1"] = 1
2. ws.cell(1,1,1)
    - ws.cell(row,column,value) 
    - 순서는 바꿔넣을경우 column=2, row=1 이런식으로 명시 필요
'''
ws["A1"]=1
ws["A2"]=2
ws["B1"]=3
ws["B2"]=4
ws.cell(3,1,5)
ws.cell(4,1,6)
ws.cell(3,2,7)
ws.cell(4,2,8)

# 셀 정보 불러오기
'''
셀 정보 불러오는법
1. ws["A1"]
2. ws.cell(1,1).value
'''
print(ws["A1"])         # A1 셀의 객체 정보
print(ws.cell(1,1))     # A1 셀의 객체 정보

print(ws["A1"].value)       # A1 셀의 값을 출력 (값이 없을경우 None 출력)
print(ws.cell(1,1).value)   # A1 셀의 값을 출력 (값이 없을경우 None 출력)


