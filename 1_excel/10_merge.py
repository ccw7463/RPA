from json import load
from openpyxl import Workbook
from openpyxl.styles import Alignment 
from openpyxl.styles import PatternFill
excel = Workbook()
ws = excel.active

# 병합하기
ws.merge_cells("B2:D2")
B2 = ws["B2"]
B2.value = "Merged Cell"
B2.alignment = Alignment(horizontal="center",vertical="center")
B2.fill = PatternFill(fgColor="FF0000",fill_type="solid")

ws.merge_cells("B4:D4")
ws.cell(4,2,"TESTING")

excel.save("merge.xlsx")

# 병합해제
from openpyxl import load_workbook
excel = load_workbook("merge.xlsx")
ws = excel.active
ws.unmerge_cells("B2:D2")
ws.unmerge_cells("B4:D4")

excel.save("unmerge.xlsx")