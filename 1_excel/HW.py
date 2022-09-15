# https://www.youtube.com/watch?v=exgO1LFl9x8&t=0s
# 영상 2:04:39

# 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
data = [[1,10,8,5,14,26,12],
        [2,7,3,7,15,24,18],
        [3,9,5,8,8,12,4],
        [4,7,8,7,17,21,18],
        [5,7,8,7,16,25,15],
        [6,3,5,8,8,17,0],
        [7,4,9,10,16,27,18],
        [8,6,6,6,15,19,17],
        [9,10,10,9,19,30,19],
        [10,9,8,8,20,25,20]]

# 데이터 기입
from openpyxl import Workbook

excel = Workbook()
ws = excel.active

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
for i in data:
    ws.append(i)

excel.save("HW.xlsx")
excel.close()

# 로드
from openpyxl import load_workbook

excel = load_workbook("HW.xlsx")
ws = excel.active


# 퀴즈2 전부 10점
cnt = 1
for i in ws["D"]:
    if cnt == 1:
        i.value = "퀴즈2"
    else:
        i.value = 10
    cnt = cnt + 1

# H열에 총점 추가
ws.cell(1,8).value = "총점"
cnt = 2
for i in range(2,12):
    # B2~G2 까지니까  
    ws.cell(i,8).value = "=SUM(" + "B" + str(cnt) + ":" + "G" + str(cnt) + ")"
    cnt = cnt + 1


# 학점 지정해주기
rank = []
for i in data:
    score = sum(i[1:3])+sum(i[4:])+10
    if score >= 90:
        rank.append("A")
    elif score >= 80:
        rank.append("B")
    elif score >= 70:
        rank.append("C")
    else:
        rank.append("D")
    
    if i[1] < 5:
        rank.pop()
        rank.append("F")

print(rank)

rank_lst = ["성적"] + rank
cnt = 0
for i in range(1,12):
    ws.cell(i,9).value = rank_lst[cnt]
    cnt = cnt + 1
excel.save("HW.xlsx")
excel.close()