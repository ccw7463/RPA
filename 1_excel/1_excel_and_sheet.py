# 설치 필요 : openpyxl
from openpyxl import Workbook

'''
@ : Workbook = EXCEL file
$ : 활성화된 sheet

◆ 메소드 정리 ◆
[1. EXCEL 기준]
@.active : 현재 활성화된 sheet
@.create_sheet() : 새로운 sheet를 기본이름으로 생성
    - @.create_sheet("MySheet1") : 새로운 sheet를 MySheet1 이름으로 생성
    - @.create_sheet("MySheet1",2) : 새로운 sheet를 3번째 자리에 생성
@.sheetnames : EXCEL 내의 시트명들 확인

@.save(이름) : EXCEL 저장
@.close() : EXCEL 닫기
@[시트명] : 시트 할당하기 
    - 시트를 딕셔너리 value 할당하는 것처럼 새로 지정 가능
        - test_sheet = excel["MySheet1"]
        - test_sheet["A1"]="테스트값입니다."
@.copy_worksheet(ws1)
    - "MySheet"이름을 가진 워크시트 객체인 ws1을 복사 --> 내용도 복사됨
    - ws1_copyed = excel.copy_worksheet(ws1)


[2. SHEET 기준]
$.title : 제목
$.sheet_properties.tabColor = RGB문자열값 : 시트이름 색상 변경
$[엑셀위치]=넣을값 
    - ws1["A1"]="Test value"
    - ws1 해당 시트에서 "A행1열"에 "Test value" 문자열 넣어줌
'''

excel = Workbook()
ws1 = excel.create_sheet() # 새로운 시트를 기본이름으로 생성
ws1.title = "MySheet1" # 시트 이름 변경
ws1.sheet_properties.tabColor = '99ff33' # RGB 형태로 값을 넣어주면 색상 변경됨(w3schools)

# 시트 생성하면서 바로 시트이름 정해주기
ws2 = excel.create_sheet("MySheet2")
ws2.sheet_properties.tabColor = "ff33cc"

# MySheet1 과 MySheet2 사이에 꼽사리시트 넣기
ws_gob = excel.create_sheet("꼽사리",2)

# 시트 이름 확인
print(excel.sheetnames)

# 시트별 작업하기
new_ws = excel["MySheet1"]
new_ws["A1"] = "Test" #  MySheet1 의 "A1" 부분에 "Test" 문자 넣기

# 시트 복사하기
ws1_copyed = excel.copy_worksheet(ws1)
ws1_copyed.title = "CopyedSheet"

# 엑셀 파일 저장                                        
excel.save("sample.xlsx")
excel.close()